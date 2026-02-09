from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Alignment

from rencons.core.dto import CertDataDTO


class UpdateHyperLinksInteractor:
    def __call__(self, wb: Workbook):
        ws = wb.active

        assert ws is not None

        for row in ws.iter_rows(min_row=2):
            link = f"{row[3].value}/{row[3].value}%{row[4].value}-{row[5].value}/{row[1].value}_{row[2].value}"  # noqa: E501

            ws[row[1].coordinate].hyperlink = link
        
        return wb


class UpdateStyleInteractor:
    def __call__(self, wb: Workbook):
        ws = wb.active

        assert ws is not None

        for e, row in enumerate(ws.iter_rows(min_row=2), start=1):
            style = "20 % - Accent1" if e % 2 else "40 % - Accent1"

            for index, cell in enumerate(row):
                cell.style = style
                alignment = Alignment(vertical="center", horizontal="center")

                if index in [2]:
                    alignment.horizontal = "left"

                assert not isinstance(cell, MergedCell)

                if any(
                    [column == cell.column_letter for column in ["K", "L", "M", "N"]]
                ):
                    alignment.wrap_text = True
                    alignment.horizontal = "left"

                cell.alignment = alignment

        return wb


class AddDataToWelderRegistryInteractor:
    def __init__(self):
        self.update_hyperlinks = UpdateHyperLinksInteractor()
        self.update_style = UpdateStyleInteractor()

    def __call__(
        self,
        registry_path: Path,
        data: list[CertDataDTO],
        group: str,
        group_key: str,
        sub_group: str,
    ):
        wb = load_workbook(registry_path)
        ws = wb.active

        assert ws is not None

        max_row = ws.max_row

        for e, el in enumerate(data, start=2):
            ws.append(
                self._data_to_ws_row(el, group, group_key, sub_group, max_row + e)
            )

        wb = self.update_hyperlinks(wb)
        wb = self.update_style(wb)

        return wb

    def _data_to_ws_row(
        self,
        data: CertDataDTO,
        group: str,
        group_key: str,
        sub_group: str,
        index: int,
    ) -> list[int | str]:
        return [
            index,
            data.kleymo,
            data.name,
            group,
            group_key,
            sub_group,
            data.cert_number,
            data.exp_date,
            data.method,
            data.gtds,
            data.thikness,
            data.outer_diameter,
            data.detail_diameter,
            data.rod_diameter,
            data.materials,
            data.detail_type,
            data.joint_type,
        ]
